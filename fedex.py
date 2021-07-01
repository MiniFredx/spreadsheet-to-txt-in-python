#! python3

# Fetches net charge and customer reference from a fedex excel sheet.

#  Imports
import pprint, openpyxl, os, time
from openpyxl.cell.read_only import EmptyCell
from openpyxl import cell
from openpyxl.utils.cell import get_column_letter
from xls2xlsx import XLS2XLSX





# TODO handle the directory to target users Documents

# Create a folder called FedexData and change the cwd to that folder.
myDir = f'{os.environ["USERPROFILE"]}\\Documents\\FedexData'
if not os.path.isdir(myDir):
    os.mkdir(myDir)
os.chdir(myDir)

# print(os.getcwd())


# Create the workspace and worksheet
x2x = XLS2XLSX("fedex.xls")
wb = x2x.to_xlsx()
# wb = openpyxl.load_workbook('fedex.xlsx')
sheet = wb.worksheets[0]

# Create 3 lists that are used for storing the different outputs given to the user
complete_data = []
incomplete_data = []
service_charge = []

# Convert the tuple given by sheet.rows to a list so we can subscript with it.
list_of_rows = list(sheet.rows)

# Get the column for each piece of data we need.
for cellObj in list_of_rows[0]:
    if(str(cellObj.value) == 'Net Charge Amount'):
        net_charge_column = cellObj.column # K Stores the column NUMBER for Net Charge
    elif(str(cellObj.value) == 'Original Customer Reference'):
        original_customer_ref = cellObj.column # AW
    elif(str(cellObj.value) == 'Original Ref#3/PO Number'):
        ref_with_po_num = cellObj.column # AY
    elif(str(cellObj.value) == 'Original Department Reference Description'):
        original_dept_ref_descr = cellObj.column # AZ
    elif(str(cellObj.value) == 'Recipient Name'):
        recipient_name = cellObj.column # AG
    elif(str(cellObj.value) == 'Recipient City'):
        recipient_city = cellObj.column # AK


# Go through each column getting the necessary info, ignoring the entire row if column a is blank
# IF all required info exists, we append it as a list to  the list complete_data. If it doesnt have the necessary data we append it as a list to the list with
# AZ (if AZ exists) + AG + AK so the information can be processed and reviewed by the user. If neither K+AW+AY or K+AG+AK exist we assume
# it must be the service charge and store it as such.

for i in range(2, sheet.max_column + 1):
        if(sheet['A'+ str(i)].value):
            ncm = sheet[get_column_letter(net_charge_column) + str(i)].value
            ocr = sheet[get_column_letter(original_customer_ref) + str(i)].value
            rwpn = sheet[get_column_letter(ref_with_po_num) + str(i)].value
            odrd = sheet[get_column_letter(original_dept_ref_descr) + str(i)].value
            rn = sheet[get_column_letter(recipient_name) + str(i)].value
            rc = sheet[get_column_letter(recipient_city) + str(i)].value
            if(ncm and ocr and rwpn):
                new_list = ['This is from row number: {}'.format(i)];
                new_list.append(f'PO: {rwpn}')
                new_list.append(f'Original Customer: {ocr}')
                new_list.append(f'Net Charge Amount: {ncm}')
                new_list.append('_' * 50)
                complete_data.append(new_list)
            elif(rn and rc):
                new_list = ['This incomplete order is from row number: {}'.format(i)];
                if(odrd):
                    new_list.append(f'Original Dept Ref: {odrd}')
                new_list.append(f'Recipient: {rn}')
                new_list.append(f'City: {rc}')
                new_list.append(f'Net Charge Amount: {ncm}')
                new_list.append('_' * 50)
                incomplete_data.append(new_list)
            else:
                new_list = ['This is the service charge from row number: {}'.format(i)]
                new_list.append(f'Service Charge Amount: {ncm}')
                service_charge.append(new_list)

def printData(list):
    str = ''
    for items in list:
        for item in items:
            str += (item) + '\n'
    return str

# Writes the data to the text file using our printData function.
# TODO if the data needs to be kept, create a naming system for the txt files
data_file = open('fedexData.txt', 'w')
data_file.write('THE FOLLOWING ORDERS WERE PRODUCED FROM COMPLETE DATA\n' + '_' * 50 + '\n' * 3)
data_file.write(printData(complete_data))
data_file.write(('\n' * 5) + (('*' * 50) + ('\n'))*3)
data_file.write('THE FOLLOWING ORDERS WERE PRODUCED FROM INCOMPLETED DATA \n\tAND NEEDS FURTHER REVIEW\n' + '_' * 50 + '\n' * 3)
data_file.write(printData(incomplete_data))
data_file.write(('\n' * 5) + (('*' * 50) + ('\n'))*3)
data_file.write('THE FOLLOWING IS THE INCLUDED SERVICE CHARGE\n' + '_' * 50 + '\n' * 3)
data_file.write(printData(service_charge))
data_file.close()
